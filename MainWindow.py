from tkinter import *
from tkinter import messagebox
import hashlib
from openpyxl import load_workbook

def encode_password(password):
    return hashlib.sha256(bytes(password, encoding = 'utf-8')).hexdigest()

class SingIn(Frame):
    def __init__(self, parent):
        Frame.__init__(self, parent, background='white')
        self.parent = parent
        # style
        self.columnconfigure(0, weight=1)
        self.columnconfigure(2, weight=1)
        self.rowconfigure(0, weight=1)
        self.rowconfigure(8, weight=1)

        self.columnconfigure(1, weight=7)
        self.rowconfigure(1, weight=5)
        self.rowconfigure(2, weight=5)
        self.rowconfigure(3, weight=5)
        self.rowconfigure(4, weight=5)
        self.rowconfigure(5, weight=5)
        self.rowconfigure(6, weight=5)
        self.rowconfigure(7, weight=5)

        self.parent = parent
        self.tries = 3
        self.lgn_lbl = Label(self, text="LOGIN", bg='white')
        self.pswrd_lbl = Label(self, text="PASSWORD", bg='white')
        self.lgn_var = StringVar()
        self.lgn_entry = Entry(self, textvariable=self.lgn_var)
        self.pswrd_var = StringVar()
        self.pswrd_entry = Entry(self, textvariable=self.pswrd_var, show="*")
        self.ok_btn = Button(self, text="CONFIRM", command=self.SignIn)
        self.ext_btn = Button(self, text="BACK", command=self.parent.show_MainWindow_menu)

        self.lgn_lbl.grid(row=1, column=1, sticky="wsen")
        self.lgn_entry.grid(row=2, column=1, sticky="wsen")
        self.pswrd_lbl.grid(row=3, column=1, sticky="wsen")
        self.pswrd_entry.grid(row=4, column=1, sticky="wsen")
        self.ok_btn.grid(row=6, column=1, sticky="wsen")
        self.ext_btn.grid(row=7, column=1, sticky="wsen")

    def SignIn(self):
        succes = 1
        status = 0

        wb = load_workbook('DB.xlsx')
        sheet = wb.active
        if self.lgn_var.get() and self.pswrd_var.get():
            i = 1
            while sheet['A' + str(i)].value:
                if str(sheet['A' + str(i)].value) == self.lgn_var.get() and str(sheet['B' + str(i)].value) == encode_password(self.pswrd_var.get()) and str(sheet.cell(i, 3).value) != 'blocked':
                    status = succes
                    break
                i += 1
            if status == 0:
                self.tries -= 1
                messagebox.showinfo("info", "Access denied, you have '{}' tries".format(self.tries))
        else:
            messagebox.showinfo("info", "Empty fields")
        if status == succes:
            self.parent.Status = self.lgn_var.get()
            self.parent.show_Main_menu()
        wb.close()
        self.lgn_var.set('')
        self.pswrd_var.set('')
        if self.tries == 0:
            sys.exit()


class SingUp(Frame):
    def __init__(self, parent):
        Frame.__init__(self, parent, background='white')
        self.parent = parent
        # style
        self.columnconfigure(0, weight=1)
        self.columnconfigure(2, weight=1)
        self.rowconfigure(0, weight=1)
        self.rowconfigure(8, weight=1)

        self.columnconfigure(1, weight=7)
        self.rowconfigure(1, weight=5)
        self.rowconfigure(2, weight=5)
        self.rowconfigure(3, weight=5)
        self.rowconfigure(4, weight=5)
        self.rowconfigure(5, weight=5)
        self.rowconfigure(6, weight=5)
        self.rowconfigure(7, weight=5)

        self.parent = parent
        self.lgn_lbl = Label(self, text="LOGIN", bg='white')
        self.pswrd_lbl = Label(self, text="PASSWORD", bg='white')
        self.lgn_var = StringVar()
        self.lgn_entry = Entry(self, textvariable=self.lgn_var)
        self.pswrd_var = StringVar()
        self.pswrd_entry = Entry(self, textvariable=self.pswrd_var, show="*")
        self.ok_btn = Button(self, text="CONFIRM", command=self.SignUp)
        self.ext_btn = Button(self, text="BACK", command=self.parent.show_MainWindow_menu)

        self.lgn_lbl.grid(row=1, column=1, sticky="wsen")
        self.lgn_entry.grid(row=2, column=1, sticky="wsen")
        self.pswrd_lbl.grid(row=3, column=1, sticky="wsen")
        self.pswrd_entry.grid(row=4, column=1, sticky="wsen")
        self.ok_btn.grid(row=6, column=1, sticky="wsen")
        self.ext_btn.grid(row=7, column=1, sticky="wsen")

    def SignUp(self):
        wb = load_workbook('DB.xlsx')
        sheet = wb.active
        if self.lgn_var.get() and self.pswrd_var.get():
            if self.parent.Limit == True and self.lgn_var.get() == self.reverse_string(self.pswrd_var.get()):
                messagebox.showinfo('info', 'Bad password')
                wb.close()
                return
            i = 1
            while sheet.cell(i, 1).value:
                if sheet.cell(i, 1).value == self.lgn_var.get():
                    messagebox.showinfo('info', 'User already exists')
                    return
                i += 1
            sheet['A' + str(i)].value = self.lgn_var.get()
            sheet['B' + str(i)].value = encode_password(self.pswrd_var.get())
            self.parent.show_MainWindow_menu()
        wb.save('DB.xlsx')
        wb.close()
        self.pswrd_var.set('')
        self.lgn_var.set('')

    def reverse_string(self, string_to_revers):
        string_to_revers = list(string_to_revers)
        complete = ''
        string_to_revers.reverse()
        for i in string_to_revers:
            complete += i
        return complete


class Users(Frame):
    def __init__(self, parent):
        Frame.__init__(self, parent, background='white')
        self.parent = parent
        self.parent.title("Users")
        # style
        self.columnconfigure(0, weight=1)
        self.columnconfigure(2, weight=1)
        self.rowconfigure(0, weight=1)
        self.rowconfigure(2, weight=1)
        self.rowconfigure(4, weight=1)

        self.columnconfigure(1, weight=7)
        self.rowconfigure(1, weight=5)
        self.rowconfigure(3, weight=5)

        self.text_field = Text(self)
        self.back_btn = Button(self, text='BACK', command=self.back)

        self.text_field.grid(row=1, column=1, sticky='wsen')
        self.back_btn.grid(row=3, column=1, sticky='wsen')
        self.show_users()

    def show_users(self):
        self.text_field.delete('1.0', END)
        wb = load_workbook('DB.xlsx')
        sheet = wb.active
        i = 1
        while sheet['A' + str(i)].value:
            self.text_field.insert(END, sheet['A' + str(i)].value + '\n')
            i += 1
        wb.close()

    def back(self):
        self.parent.show_Main_menu()


class AddUser(Frame):
    def __init__(self, parent):
        Frame.__init__(self, parent, background='white')
        self.parent = parent
        self.parent.title("Add user")
        # style
        self.columnconfigure(0, weight=1)
        self.columnconfigure(2, weight=1)
        self.rowconfigure(0, weight=1)
        self.rowconfigure(3, weight=1)
        self.rowconfigure(6, weight=1)

        self.columnconfigure(1, weight=7)
        self.rowconfigure(1, weight=5)
        self.rowconfigure(2, weight=5)
        self.rowconfigure(4, weight=5)
        self.rowconfigure(5, weight=5)

        self.user_name_label = Label(self, text='USER NAME', bg='white')
        self.user_name_var = StringVar()
        self.user_name_entry = Entry(self, textvariable=self.user_name_var)
        self.apply_btn = Button(self, text='APPLY', command=self.apply)
        self.back_btn = Button(self, text='BACK', command=self.back)

        self.user_name_label.grid(row=1, column=1, sticky='wsen')
        self.user_name_entry.grid(row=2, column=1, sticky='wsen')
        self.apply_btn.grid(row=4, column=1, sticky='wsen')
        self.back_btn.grid(row=5, column=1, sticky='wsen')

    def apply(self):
        if self.user_name_var.get():
            wb = load_workbook('DB.xlsx')
            sheet = wb.active
            i = 1
            while sheet['A' + str(i)].value:
                if sheet['A' + str(i)].value == self.user_name_var.get():
                    messagebox.showinfo("info", "User is already in DB")
                    return
                i += 1
            sheet['A' + str(i)].value = self.user_name_var.get()
            wb.save('DB.xlsx')
            wb.close()
            messagebox.showinfo("info", "Success")
        else:
            messagebox.showinfo("info", "Empty field")

    def back(self):
        self.parent.show_Main_menu()


class BlockUnblockUser(Frame):
    def __init__(self, parent):
        Frame.__init__(self, parent, background='white')
        self.parent = parent
        self.parent.title("Block/Unblock user")
        # style
        self.columnconfigure(0, weight=1)
        self.columnconfigure(2, weight=1)
        self.rowconfigure(0, weight=1)
        self.rowconfigure(3, weight=1)
        self.rowconfigure(6, weight=1)

        self.columnconfigure(1, weight=7)
        self.rowconfigure(1, weight=5)
        self.rowconfigure(2, weight=5)
        self.rowconfigure(4, weight=5)
        self.rowconfigure(5, weight=5)

        self.user_name_label = Label(self, text='USER NAME', bg='white')
        self.user_name_var = StringVar()
        self.user_name_entry = Entry(self, textvariable=self.user_name_var)
        self.apply_btn = Button(self, text='APPLY', command=self.apply)
        self.back_btn = Button(self, text='BACK', command=self.back)

        self.user_name_label.grid(row=1, column=1, sticky='wsen')
        self.user_name_entry.grid(row=2, column=1, sticky='wsen')
        self.apply_btn.grid(row=4, column=1, sticky='wsen')
        self.back_btn.grid(row=5, column=1, sticky='wsen')

    def apply(self):
        if self.user_name_var.get():
            wb = load_workbook('DB.xlsx')
            sheet = wb.active
            i = 1
            found = False
            while sheet['A' + str(i)].value:
                if sheet['A' + str(i)].value == self.user_name_var.get():
                    if sheet['C' + str(i)].value == 'blocked':
                        sheet['C' + str(i)].value = ''
                    else:
                        sheet['C' + str(i)].value = 'blocked'
                    messagebox.showinfo("info", "Success")
                    wb.save('DB.xlsx')
                    found = True
                    break
                i += 1
            wb.close()
            if not found:
                messagebox.showinfo("info", "User not found")
        else:
            messagebox.showinfo("info", "Empty field")

    def back(self):
        self.parent.show_Main_menu()


class MainMenu(Frame):
    def __init__(self, parent, status):
        Frame.__init__(self, parent, background='white')
        self.parent = parent
        self.parent.title("Main Window")
        self.Status = status
        # style
        self.columnconfigure(0, weight=1)
        self.columnconfigure(2, weight=1)
        self.rowconfigure(0, weight=1)
        self.rowconfigure(6, weight=1)
        self.rowconfigure(8, weight=1)

        self.columnconfigure(1, weight=7)
        self.rowconfigure(1, weight=5)
        self.rowconfigure(2, weight=5)
        self.rowconfigure(3, weight=5)
        self.rowconfigure(4, weight=5)
        self.rowconfigure(5, weight=5)
        self.rowconfigure(7, weight=5)

        self.change_pswrd_btn = Button(self, text='CHANGE PASSWORD', command=self.change_password)
        self.show_all_users_btn = Button(self, text='SHOW ALL USERS', command=self.show_all_users)
        self.add_new_user_btn = Button(self, text='ADD NEW USER', command=self.add_new_user)
        self.block_user_btn = Button(self, text='BLOCK/UNBLOCK USER', command=self.block_user)
        self.add_remove_limits_btn = Button(self, text='ADD/REMOVE LIMITS', command=self.add_remove_limits)
        self.sign_out_btn = Button(self, text='SIGN OUT', command=self.sign_out)

        self.change_pswrd_btn.grid(row=1, column=1, sticky='wsen')
        self.show_all_users_btn.grid(row=2, column=1, sticky='wsen')
        self.add_new_user_btn.grid(row=3, column=1, sticky='wsen')
        self.block_user_btn.grid(row=4, column=1, sticky='wsen')
        self.add_remove_limits_btn.grid(row=5, column=1, sticky='wsen')
        self.sign_out_btn.grid(row=7, column=1, sticky='wsen')

    def change_password(self):
        self.parent.show_NewPassword_menu()

    def show_all_users(self):
        if self.parent.Status == 'Admin':
            self.parent.show_Users_menu()
        else:
            messagebox.showinfo('info', 'Access denied')

    def add_new_user(self):
        if self.parent.Status == 'Admin':
            self.parent.show_AddUser_menu()
        else:
            messagebox.showinfo('info', 'Access denied')

    def block_user(self):
        if self.parent.Status == 'Admin':
            self.parent.show_BlockUnblockUser_menu()
        else:
            messagebox.showinfo('info', 'Access denied')

    def add_remove_limits(self):
        if self.parent.Limit:
            self.parent.Limit = False
        else:
            self.parent.Limit = True

    def sign_out(self):
        self.parent.Status = 0
        self.parent.show_MainWindow_menu()


class NewPassword(Frame):
    def __init__(self, parent):
        Frame.__init__(self, parent, background='white')
        self.parent = parent
        self.parent.title("New Password")
        # style
        self.columnconfigure(0, weight=1)
        self.columnconfigure(2, weight=1)
        self.rowconfigure(0, weight=1)
        self.rowconfigure(7, weight=1)
        self.rowconfigure(10, weight=1)

        self.columnconfigure(1, weight=7)
        self.rowconfigure(1, weight=5)
        self.rowconfigure(2, weight=5)
        self.rowconfigure(3, weight=5)
        self.rowconfigure(4, weight=5)
        self.rowconfigure(5, weight=5)
        self.rowconfigure(6, weight=5)
        self.rowconfigure(8, weight=5)
        self.rowconfigure(9, weight=5)

        self.label_old_pswrd = Label(self, text='Old password', bg='white')
        self.old_password_var = StringVar()
        self.old_password_entry = Entry(self, textvariable=self.old_password_var, show="*")
        self.label_new_pswrd = Label(self, text='New password', bg='white')
        self.new_password_var = StringVar()
        self.new_password_entry = Entry(self, textvariable=self.new_password_var, show="*")
        self.label_new_again_pswrd = Label(self, text='New password once more', bg='white')
        self.new_again_password_var = StringVar()
        self.new_again_password_entry = Entry(self, textvariable=self.new_again_password_var, show="*")
        self.apply_btn = Button(self, text='APPLY', command=self.apply)
        self.back_btn = Button(self, text='BACK', command=self.back)

        self.label_old_pswrd.grid(row=1, column=1, sticky="wsen")
        self.old_password_entry.grid(row=2, column=1, sticky="wsen")
        self.label_new_pswrd.grid(row=3, column=1, sticky="wsen")
        self.new_password_entry.grid(row=4, column=1, sticky="wsen")
        self.label_new_again_pswrd.grid(row=5, column=1, sticky="wsen")
        self.new_again_password_entry.grid(row=6, column=1, sticky="wsen")
        self.apply_btn.grid(row=8, column=1, sticky="wsen")
        self.back_btn.grid(row=9, column=1, sticky="wsen")

    def apply(self):
        wb = load_workbook('DB.xlsx')
        sheet = wb.active
        i = 1
        while sheet['A' + str(i)].value != self.parent.Status:
            i += 1
        if encode_password(self.old_password_var.get()) == sheet['B' + str(i)].value:
            if self.new_again_password_var.get() == self.new_password_var.get():
                sheet['B' + str(i)].value = encode_password(self.new_password_var.get())
                wb.save('DB.xlsx')
        wb.close()

    def back(self):
        self.parent.show_Main_menu()


class MainWindow(Frame):
    def __init__(self, parent):
        Frame.__init__(self, parent, background='white')
        self.parent = parent
        self.parent.title("Main Window")
        # style
        self.columnconfigure(0, weight=1)
        self.columnconfigure(2, weight=1)
        self.rowconfigure(0, weight=1)
        self.rowconfigure(4, weight=1)

        self.columnconfigure(1, weight=7)
        self.rowconfigure(1, weight=5)
        self.rowconfigure(2, weight=5)
        self.rowconfigure(3, weight=5)

        self.signIn_btn = Button(self, text="SIGN IN", command=self.parent.show_SignIn_menu)
        self.signUp_btn = Button(self, text="SIGN UP", command=self.parent.show_SignUp_menu)
        self.extBtn = Button(self, text="EXIT", command=sys.exit)
        self.signIn_btn.grid(row=1, column=1, sticky="wsen")
        self.signUp_btn.grid(row=2, column=1, sticky="wsen")
        self.extBtn.grid(row=3, column=1, sticky="wsen")


class MyProgram(Tk):
    def __init__(self):
        Tk.__init__(self)
        self.title("Main Window")
        self.geometry("300x400")
        self.Status = 0
        self.Limit = False
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)
        self.AddUser_menu = AddUser(self)
        self.BlockUnblockUser_menu = BlockUnblockUser(self)
        self.Main_menu = MainMenu(self, self.Status)
        self.MainWindow_menu = MainWindow(self)
        self.NewPassword_menu = NewPassword(self)
        self.SignIn_menu = SingIn(self)
        self.SignUp_menu = SingUp(self)
        self.Users_menu = Users(self)

    def show_AddUser_menu(self):
        self.geometry("300x400")
        self.BlockUnblockUser_menu.grid_forget()
        self.Main_menu.grid_forget()
        self.NewPassword_menu.grid_forget()
        self.SignIn_menu.grid_forget()
        self.SignUp_menu.grid_forget()
        self.Users_menu.grid_forget()
        self.MainWindow_menu.grid_forget()
        self.AddUser_menu.grid(row=0, column=0, sticky="wsen")

    def show_BlockUnblockUser_menu(self):
        self.geometry("300x400")
        self.BlockUnblockUser_menu.grid(row=0, column=0, sticky="wsen")
        self.Main_menu.grid_forget()
        self.NewPassword_menu.grid_forget()
        self.SignIn_menu.grid_forget()
        self.SignUp_menu.grid_forget()
        self.Users_menu.grid_forget()
        self.MainWindow_menu.grid_forget()
        self.AddUser_menu.grid_forget()

    def show_Main_menu(self):
        self.geometry("500x600")
        self.BlockUnblockUser_menu.grid_forget()
        self.Main_menu.grid(row=0, column=0, sticky="wsen")
        self.NewPassword_menu.grid_forget()
        self.SignIn_menu.grid_forget()
        self.SignUp_menu.grid_forget()
        self.Users_menu.grid_forget()
        self.MainWindow_menu.grid_forget()
        self.AddUser_menu.grid_forget()

    def show_MainWindow_menu(self):
        self.geometry("300x400")
        self.BlockUnblockUser_menu.grid_forget()
        self.Main_menu.grid_forget()
        self.NewPassword_menu.grid_forget()
        self.SignIn_menu.grid_forget()
        self.SignUp_menu.grid_forget()
        self.MainWindow_menu.grid(row=0, column=0, sticky="wsen")
        self.Users_menu.grid_forget()
        self.AddUser_menu.grid_forget()

    def show_NewPassword_menu(self):
        self.geometry("300x400")
        self.BlockUnblockUser_menu.grid_forget()
        self.Main_menu.grid_forget()
        self.NewPassword_menu.grid(row=0, column=0, sticky="wsen")
        self.SignIn_menu.grid_forget()
        self.SignUp_menu.grid_forget()
        self.Users_menu.grid_forget()
        self.MainWindow_menu.grid_forget()
        self.AddUser_menu.grid_forget()

    def show_SignIn_menu(self):
        self.geometry("300x400")
        self.BlockUnblockUser_menu.grid_forget()
        self.Main_menu.grid_forget()
        self.NewPassword_menu.grid_forget()
        self.SignIn_menu.grid(row=0, column=0, sticky="wsen")
        self.SignUp_menu.grid_forget()
        self.MainWindow_menu.grid_forget()
        self.Users_menu.grid_forget()
        self.AddUser_menu.grid_forget()

    def show_SignUp_menu(self):
        self.geometry("300x400")
        self.BlockUnblockUser_menu.grid_forget()
        self.Main_menu.grid_forget()
        self.NewPassword_menu.grid_forget()
        self.SignIn_menu.grid_forget()
        self.SignUp_menu.grid(row=0, column=0, sticky="wsen")
        self.Users_menu.grid_forget()
        self.AddUser_menu.grid_forget()

    def show_Users_menu(self):
        self.geometry("800x600")
        self.BlockUnblockUser_menu.grid_forget()
        self.Main_menu.grid_forget()
        self.NewPassword_menu.grid_forget()
        self.SignIn_menu.grid_forget()
        self.SignUp_menu.grid_forget()
        self.Users_menu.grid(row=0, column=0, sticky="wsen")
        self.Users_menu.show_users()
        self.MainWindow_menu.grid_forget()
        self.AddUser_menu.grid_forget()


def check():
    try:
        license_file = open("F:/license.txt", "r")
        if license_file.read() == "1234":
            root = MyProgram()
            root.show_MainWindow_menu()
            root.mainloop()
        else:
            messagebox.showinfo("info", "License error")
    except FileNotFoundError:
        messagebox.showinfo("info", "License error")


def main():
    check()


if __name__ == '__main__':
    main()
